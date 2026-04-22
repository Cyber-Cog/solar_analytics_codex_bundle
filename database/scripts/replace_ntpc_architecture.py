"""
Replace NTPCNOKHRA plant_architecture with data from the CORRECT file:
  Corrected_AsBuilt_SCB_Mapping (1).xlsx

Column mapping (from row 2 headers in that file):
  inverter_id       -> Col A  (raw: '01-A'   -> DB: 'INV-01A')
  scb_id            -> Col B  (raw: '01-A-01' -> DB: 'SCB-01A-01')
  Inverter_make     -> Col C
  Module_make       -> Col D
  Module_Technology -> Col E
  modules_per_string-> Col F  (constant 30)
  Module_watt       -> Col G  (varies: 330/335/340/345/350)
  strings_per_scb   -> Col H  (VARIES: 23-30, critical for DS algo)
  dc_capacity_kw    -> Col I  (total SCB kW = strings x modules x watt/1000)

Safe to re-run (DELETE + COPY).  Does NOT touch raw_data_generic or fault_diagnostics.
"""
import io
import sys
import openpyxl
import pandas as pd
import psycopg2

EXCEL_PATH = r"C:\Users\Asus\Downloads\Corrected_AsBuilt_SCB_Mapping (1).xlsx"
PG_DSN     = "postgresql://solar:solar@localhost:5432/solar"
PLANT_ID   = "NTPCNOKHRA"


# ── ID converters ────────────────────────────────────────────────────────────
def to_db_inv(raw: str) -> str:
    """'01-A' -> 'INV-01A'"""
    parts = str(raw).strip().split("-")  # ['01', 'A']
    return "INV-" + "".join(parts)


def to_db_scb(raw: str) -> str:
    """'01-A-01' -> 'SCB-01A-01'"""
    parts = str(raw).strip().split("-")  # ['01', 'A', '01']
    return "SCB-" + parts[0] + parts[1] + "-" + parts[2]


def to_str_id(scb_raw: str, str_num: int) -> str:
    """'01-A-01', 3 -> 'STR-01A-01-03'"""
    parts = str(scb_raw).strip().split("-")
    return f"STR-{parts[0]}{parts[1]}-{parts[2]}-{str_num:02d}"


# ── Load Excel ───────────────────────────────────────────────────────────────
def load_excel() -> pd.DataFrame:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # Row 1 = title row, Row 2 = column headers, Row 3+ = data
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    print(f"  Excel headers: {headers}")

    rows = []
    for r in range(3, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value
               for c in range(1, ws.max_column + 1)}
        rows.append(row)

    df = pd.DataFrame(rows)

    # Convert IDs to DB format
    df["inverter_id_db"] = df["inverter_id"].apply(to_db_inv)
    df["scb_id_db"]      = df["scb_id"].apply(to_db_scb)

    # Numeric coercion
    df["strings_per_scb"]    = pd.to_numeric(df["strings_per_scb"], errors="coerce").astype(int)
    df["modules_per_string"] = pd.to_numeric(df["modules_per_string"], errors="coerce").astype(int)
    df["Module_watt"]        = pd.to_numeric(df["Module_watt"], errors="coerce").astype(float)

    # Per-string DC capacity (kW) = modules_per_string × module_watt / 1000
    df["dc_cap_per_string_kw"] = df["modules_per_string"] * df["Module_watt"] / 1000.0

    # SCBs per inverter (computed from actual data)
    df["scbs_per_inverter"] = df.groupby("inverter_id_db")["scb_id_db"].transform("count").astype(int)

    return df


# ── Build string-level rows ──────────────────────────────────────────────────
def build_rows(df: pd.DataFrame):
    rows = []
    for _, scb in df.iterrows():
        n = int(scb["strings_per_scb"])
        for s in range(1, n + 1):
            rows.append((
                PLANT_ID,
                scb["inverter_id_db"],
                scb["scb_id_db"],
                to_str_id(str(scb["scb_id"]), s),
                int(scb["modules_per_string"]),
                n,
                int(scb["scbs_per_inverter"]),
                float(scb["dc_cap_per_string_kw"]),
            ))
    return rows


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print(f"Loading: {EXCEL_PATH}")
    df = load_excel()
    print(f"  SCBs: {len(df)} | Inverters: {df['inverter_id_db'].nunique()}")
    print(f"  strings_per_scb range: {df['strings_per_scb'].min()}–{df['strings_per_scb'].max()}")
    print(f"  modules_per_string unique: {sorted(df['modules_per_string'].unique())}")
    print(f"  Module_watt unique: {sorted(df['Module_watt'].unique())}")

    rows = build_rows(df)
    total_str_rows = len(rows)
    print(f"  Total string rows to insert: {total_str_rows}")

    conn = psycopg2.connect(PG_DSN)
    conn.autocommit = False
    cur = conn.cursor()

    try:
        # Sanity check existing
        cur.execute("SELECT COUNT(*) FROM plant_architecture WHERE plant_id=%s", (PLANT_ID,))
        existing = cur.fetchone()[0]
        print(f"\nExisting rows for {PLANT_ID}: {existing}")

        # Fix sequence first (avoid duplicate-key issues)
        cur.execute("SELECT MAX(id) FROM plant_architecture")
        max_id = cur.fetchone()[0] or 0
        cur.execute(f"SELECT setval('plant_architecture_id_seq', {max(max_id + 1, total_str_rows + 100000)}, false)")

        # Delete old data
        print(f"Deleting existing {PLANT_ID} architecture rows...")
        cur.execute("DELETE FROM plant_architecture WHERE plant_id=%s", (PLANT_ID,))
        print(f"  Deleted {cur.rowcount} rows.")

        # COPY new rows
        print("Inserting via COPY...")
        buf = io.StringIO()
        for (plant_id, inv_id, scb_id, str_id, mps, sps, spi, dc_cap) in rows:
            buf.write(f"{plant_id}\t{inv_id}\t{scb_id}\t{str_id}\t{mps}\t{sps}\t{spi}\t{dc_cap}\n")
        buf.seek(0)
        cur.copy_from(
            buf, "plant_architecture",
            columns=("plant_id", "inverter_id", "scb_id", "string_id",
                     "modules_per_string", "strings_per_scb",
                     "scbs_per_inverter", "dc_capacity_kw"),
        )
        print(f"  Inserted {cur.rowcount} rows.")

        # Clear fault_cache for this plant
        cur.execute("DELETE FROM fault_cache WHERE cache_key LIKE %s", (f"%{PLANT_ID}%",))
        print(f"  Cleared fault_cache rows: {cur.rowcount}")

        conn.commit()
        print("\nCommitted successfully.")

        # Verify
        cur.execute("SELECT COUNT(*) FROM plant_architecture WHERE plant_id=%s", (PLANT_ID,))
        print(f"New total rows for {PLANT_ID}: {cur.fetchone()[0]}")

        # Quick spot check
        cur.execute("""
            SELECT scb_id, strings_per_scb, modules_per_string, dc_capacity_kw
            FROM plant_architecture WHERE plant_id=%s
            ORDER BY inverter_id, scb_id, string_id LIMIT 6
        """, (PLANT_ID,))
        cols = [d[0] for d in cur.description]
        print("\nSample rows:")
        for r in cur.fetchall():
            print("  ", dict(zip(cols, r)))

        # Distribution of strings_per_scb
        cur.execute("""
            SELECT strings_per_scb, COUNT(DISTINCT scb_id) AS scbs
            FROM plant_architecture WHERE plant_id=%s
            GROUP BY strings_per_scb ORDER BY strings_per_scb
        """, (PLANT_ID,))
        print("\nstrings_per_scb distribution (distinct SCBs):")
        for r in cur.fetchall():
            print(f"  {r[0]} strings → {r[1]} SCBs")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR (rolled back): {e}")
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
